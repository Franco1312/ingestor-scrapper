"""
Health checks - Generic validation functions for content monitoring.

This module provides pure functions for validating fetched content,
checking schemas, and detecting changes.
"""

import csv
import hashlib
import logging
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# Try to import optional libraries
try:
    from bs4 import BeautifulSoup  # type: ignore

    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False
    logger.warning(
        "BeautifulSoup4 not installed. HTML selector checks will use fallback."
    )

try:
    import openpyxl  # type: ignore

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning(
        "openpyxl not installed. Excel schema checks will be skipped."
    )


def fetch(
    url: str, timeout: int = 30, max_retries: int = 2, verify_ssl: bool = True
) -> Tuple[bytes, Dict[str, str], int, str]:
    """
    Fetch content from URL using requests.

    Args:
        url: URL to fetch
        timeout: Request timeout in seconds
        max_retries: Maximum number of retry attempts
        verify_ssl: Whether to verify SSL certificates (default: True)

    Returns:
        Tuple of (content_bytes, headers_dict, status_code, final_url)

    Raises:
        ImportError: If requests library not installed
        Exception: On network/HTTP errors
    """
    try:
        import requests  # type: ignore
    except ImportError:
        raise ImportError(
            "requests library required for health checks. "
            "Install with: pip install requests"
        )

    headers_dict: Dict[str, str] = {}
    last_exception = None

    for attempt in range(max_retries + 1):
        try:
            response = requests.get(
                url, timeout=timeout, allow_redirects=True, verify=verify_ssl
            )
            # Convert headers to dict
            headers_dict = dict(response.headers)
            return (
                response.content,
                headers_dict,
                response.status_code,
                response.url,
            )
        except requests.exceptions.RequestException as e:
            last_exception = e
            if attempt < max_retries:
                logger.warning(
                    "Request failed (attempt %d/%d): %s. Retrying...",
                    attempt + 1,
                    max_retries + 1,
                    e,
                )
                continue
            else:
                logger.error(
                    "Request failed after %d attempts: %s",
                    max_retries + 1,
                    e,
                )
                raise

    # Should never reach here, but satisfy type checker
    if last_exception:
        raise last_exception
    raise RuntimeError("Unexpected error in fetch")


def check_status(status_code: int) -> bool:
    """
    Check if HTTP status code is successful.

    Args:
        status_code: HTTP status code

    Returns:
        True if status code is 200-299
    """
    return 200 <= status_code < 300


def check_min_bytes(content: bytes, min_bytes: int) -> bool:
    """
    Check if content meets minimum size requirement.

    Args:
        content: Content bytes
        min_bytes: Minimum required bytes

    Returns:
        True if content size >= min_bytes
    """
    return len(content) >= min_bytes


def check_content_type(headers: Dict[str, str], expected: str) -> bool:
    """
    Check if Content-Type header matches expected value.

    Args:
        headers: Response headers dict
        expected: Expected Content-Type value (substring match)

    Returns:
        True if Content-Type contains expected value
    """
    content_type = headers.get("Content-Type", "").lower()
    return expected.lower() in content_type


def check_html_contains(content: bytes, selectors: List[str]) -> Dict[str, bool]:
    """
    Check if HTML content contains specified selectors.

    Uses BeautifulSoup if available, otherwise falls back to string matching.

    Args:
        content: HTML content bytes
        selectors: List of CSS selectors or strings to search

    Returns:
        Dict mapping selector -> bool indicating if found
    """
    result: Dict[str, bool] = {}

    if not selectors:
        return result

    try:
        html_text = content.decode("utf-8", errors="ignore")
    except Exception as e:
        logger.error("Failed to decode HTML: %s", e)
        return {selector: False for selector in selectors}

    if HAS_BS4:
        # Use BeautifulSoup for proper CSS selector matching
        try:
            soup = BeautifulSoup(html_text, "html.parser")
            for selector in selectors:
                try:
                    found = soup.select_one(selector) is not None
                    result[selector] = found
                except Exception:
                    # Invalid selector or parse error
                    result[selector] = False
        except Exception as e:
            logger.warning("BeautifulSoup parsing error: %s. Using fallback.", e)
            # Fall back to string matching
            for selector in selectors:
                result[selector] = selector in html_text
    else:
        # Fallback: simple string matching
        for selector in selectors:
            result[selector] = selector in html_text

    return result


def check_csv_schema(
    content: bytes, expected_columns: List[str], min_rows: Optional[int] = None
) -> Dict[str, Any]:
    """
    Check CSV schema against expected columns and minimum rows.

    Args:
        content: CSV content bytes
        expected_columns: Expected column names
        min_rows: Minimum required rows (excluding header)

    Returns:
        Dict with:
            - valid: bool
            - found_columns: List[str]
            - missing_columns: List[str]
            - row_count: int
            - row_count_valid: bool
            - error: Optional[str]
    """
    result: Dict[str, Any] = {
        "valid": False,
        "found_columns": [],
        "missing_columns": [],
        "row_count": 0,
        "row_count_valid": True,
        "error": None,
    }

    if not expected_columns and min_rows is None:
        result["valid"] = True
        return result

    try:
        text = content.decode("utf-8", errors="replace")
    except Exception as e:
        result["error"] = f"Failed to decode CSV: {e}"
        return result

    try:
        # Detect delimiter
        sniffer = csv.Sniffer()
        sample = text[:1024]  # Sample for delimiter detection
        delimiter = sniffer.sniff(sample).delimiter
    except Exception:
        # Fall back to comma
        delimiter = ","

    try:
        reader = csv.DictReader(StringIO(text), delimiter=delimiter)
        found_columns = reader.fieldnames or []
        rows = list(reader)

        result["found_columns"] = list(found_columns)
        result["row_count"] = len(rows)

        # Check expected columns
        if expected_columns:
            missing = set(expected_columns) - set(found_columns)
            result["missing_columns"] = list(missing)
            if not missing:
                result["valid"] = True
        else:
            result["valid"] = True

        # Check min rows
        if min_rows is not None:
            result["row_count_valid"] = result["row_count"] >= min_rows
            if not result["row_count_valid"]:
                result["valid"] = False
                result[
                    "error"
                ] = f"Expected at least {min_rows} rows, got {result['row_count']}"

    except Exception as e:
        result["error"] = f"CSV parsing error: {e}"
        logger.error("CSV parsing failed: %s", e, exc_info=True)

    return result


def check_excel_schema(
    content: bytes, expected_columns: List[str], min_rows: Optional[int] = None
) -> Dict[str, Any]:
    """
    Check Excel schema against expected columns and minimum rows.

    Args:
        content: Excel content bytes
        expected_columns: Expected column names
        min_rows: Minimum required rows (excluding header)

    Returns:
        Dict with:
            - valid: bool
            - found_columns: List[str]
            - missing_columns: List[str]
            - row_count: int
            - row_count_valid: bool
            - error: Optional[str]
            - skipped: bool (if openpyxl not available)

    Note: Checks first sheet only.
    """
    result: Dict[str, Any] = {
        "valid": False,
        "found_columns": [],
        "missing_columns": [],
        "row_count": 0,
        "row_count_valid": True,
        "error": None,
        "skipped": False,
    }

    if not HAS_OPENPYXL:
        result["skipped"] = True
        result["error"] = "openpyxl not installed - Excel check skipped"
        return result

    if not expected_columns and min_rows is None:
        result["valid"] = True
        return result

    try:
        workbook = openpyxl.load_workbook(BytesIO(content))
        # Use first sheet
        sheet = workbook.active

        # Read header row
        found_columns: List[str] = []
        if sheet.max_row > 0:
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[
                0
            ]
            found_columns = [str(cell) if cell is not None else "" for cell in header_row]

        result["found_columns"] = found_columns
        result["row_count"] = sheet.max_row - 1  # Exclude header

        # Check expected columns
        if expected_columns:
            missing = set(expected_columns) - set(found_columns)
            result["missing_columns"] = list(missing)
            if not missing:
                result["valid"] = True
        else:
            result["valid"] = True

        # Check min rows
        if min_rows is not None:
            result["row_count_valid"] = result["row_count"] >= min_rows
            if not result["row_count_valid"]:
                result["valid"] = False
                result[
                    "error"
                ] = f"Expected at least {min_rows} rows, got {result['row_count']}"

    except Exception as e:
        result["error"] = f"Excel parsing error: {e}"
        logger.error("Excel parsing failed: %s", e, exc_info=True)

    return result


def checksum_sha256(content: bytes) -> str:
    """
    Calculate SHA-256 checksum of content.

    Args:
        content: Content bytes

    Returns:
        Hexadecimal SHA-256 checksum string
    """
    return hashlib.sha256(content).hexdigest()

