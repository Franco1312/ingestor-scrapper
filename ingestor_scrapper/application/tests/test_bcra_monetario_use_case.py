"""
Tests for BCRA Monetario Use Case.

Tests the orchestration logic that coordinates fetching, parsing,
normalization, and output.
"""

from datetime import datetime
from io import BytesIO
from unittest.mock import Mock

import pytest  # type: ignore

from ingestor_scrapper.adapters.normalizers.bcra_monetario import (
    AdapterBcraMonetarioNormalizer,
)
from ingestor_scrapper.adapters.parsers.bcra_excel import (
    AdapterBcraExcelParser,
)
from ingestor_scrapper.application.bcra_monetario_use_case import (
    BcraMonetarioUseCase,
)
from ingestor_scrapper.core.entities import ContentType, Document


class TestBcraMonetarioUseCase:
    """Test suite for BcraMonetarioUseCase."""

    @pytest.fixture
    def mock_fetcher(self):
        """Create a mock document fetcher."""
        return Mock()

    @pytest.fixture
    def parser(self):
        """Create a real parser for testing."""
        return AdapterBcraExcelParser()

    @pytest.fixture
    def normalizer(self):
        """Create a real normalizer for testing."""
        return AdapterBcraMonetarioNormalizer()

    @pytest.fixture
    def mock_output(self):
        """Create a mock output port."""
        return Mock()

    @pytest.fixture
    def use_case(self, mock_fetcher, parser, normalizer, mock_output):
        """Create a use case with dependencies."""
        return BcraMonetarioUseCase(
            fetcher=mock_fetcher,
            parser=parser,
            normalizer=normalizer,
            output=mock_output,
        )

    def test_execute_empty_url(self, use_case):
        """Test that empty URL returns empty list."""
        items = use_case.execute("")
        assert items == []

    def test_execute_successful_flow(
        self, use_case, mock_fetcher, mock_output
    ):
        """Test successful execution of the use case."""
        # Create sample Excel bytes
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "RESERVAS"

        # Add test data
        dates = [datetime(2025, 10, 28)]
        values = [40771.0]
        for idx, (d, v) in enumerate(zip(dates, values), start=1):
            ws.cell(row=idx, column=1, value=d)
            ws.cell(row=idx, column=3, value=v)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.getvalue()

        # Mock fetcher
        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )
        mock_fetcher.fetch.return_value = document

        # Execute
        items = use_case.execute("https://example.com/test.xlsm")

        # Assert
        assert len(items) >= 1
        mock_fetcher.fetch.assert_called_once()
        mock_output.emit.assert_called_once()

    def test_execute_fetch_error(self, use_case, mock_fetcher):
        """Test handling of fetch errors."""
        mock_fetcher.fetch.side_effect = Exception("Network error")

        items = use_case.execute("https://example.com/test.xlsm")

        assert items == []
        mock_fetcher.fetch.assert_called_once()

    def test_execute_empty_document(self, use_case, mock_fetcher):
        """Test handling of empty document."""
        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=b"",
        )
        mock_fetcher.fetch.return_value = document

        items = use_case.execute("https://example.com/test.xlsm")

        assert items == []

    def test_execute_parse_error(self, use_case, mock_fetcher):
        """Test handling of parse errors."""
        # Create an invalid Excel
        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=b"invalid excel content",
        )
        mock_fetcher.fetch.return_value = document

        items = use_case.execute("https://example.com/test.xlsm")

        assert items == []

    def test_execute_no_records_extracted(self, mock_fetcher):
        """Test handling when no records are extracted."""
        # Mock a parser that returns empty list
        mock_parser = Mock()
        mock_parser.parse.return_value = []

        use_case_no_records = BcraMonetarioUseCase(
            fetcher=mock_fetcher,
            parser=mock_parser,
            normalizer=self._get_normalizer(),
            output=Mock(),
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=b"dummy",
        )
        mock_fetcher.fetch.return_value = document

        items = use_case_no_records.execute("https://example.com/test.xlsm")

        assert items == []

    def test_execute_normalization_error(self, use_case, mock_fetcher):
        """Test handling of normalization errors."""
        # Create valid Excel but mock normalizer that fails
        mock_normalizer = Mock()
        mock_normalizer.normalize.side_effect = Exception(
            "Normalization error"
        )

        use_case_fail_norm = BcraMonetarioUseCase(
            fetcher=mock_fetcher,
            parser=self._get_parser(),
            normalizer=mock_normalizer,
            output=Mock(),
        )

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "RESERVAS"
        ws.cell(row=1, column=1, value=datetime(2025, 10, 28))
        ws.cell(row=1, column=3, value=40771.0)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )
        mock_fetcher.fetch.return_value = document

        items = use_case_fail_norm.execute("https://example.com/test.xlsm")

        assert items == []

    def test_execute_output_error_does_not_fail(
        self, mock_fetcher, parser, normalizer
    ):
        """Test that output errors don't fail the execution."""
        # Create mock output that fails
        mock_output_fail = Mock()
        mock_output_fail.emit.side_effect = Exception("Output error")

        use_case = BcraMonetarioUseCase(
            fetcher=mock_fetcher,
            parser=parser,
            normalizer=normalizer,
            output=mock_output_fail,
        )

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "RESERVAS"
        ws.cell(row=1, column=1, value=datetime(2025, 10, 28))
        ws.cell(row=1, column=3, value=40771.0)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )
        mock_fetcher.fetch.return_value = document

        # Should not raise, just return items
        items = use_case.execute("https://example.com/test.xlsm")

        assert len(items) >= 1

    @staticmethod
    def _get_parser():
        """Helper to get a parser."""
        return AdapterBcraExcelParser()

    @staticmethod
    def _get_normalizer():
        """Helper to get a normalizer."""
        return AdapterBcraMonetarioNormalizer()
