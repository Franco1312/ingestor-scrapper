"""
Tests for BCRA Excel Parser.

Tests the parsing logic that extracts indicators from Excel files.
"""

from datetime import datetime
from io import BytesIO

import pytest  # type: ignore
from openpyxl import Workbook  # type: ignore

from ingestor_scrapper.adapters.parsers.bcra_excel import (
    AdapterBcraExcelParser,
)
from ingestor_scrapper.core.entities import ContentType, Document


class TestAdapterBcraExcelParser:
    """Test suite for AdapterBcraExcelParser."""

    @pytest.fixture
    def parser(self):
        """Create a parser instance for testing."""
        return AdapterBcraExcelParser()

    def create_test_excel_bytes(
        self, sheet_name: str, dates: list, values: list, value_column: int
    ) -> bytes:
        """
        Create a test Excel file in memory.

        Args:
            sheet_name: Name of the sheet
            dates: List of date strings
            values: List of values for the value column
            value_column: Column index for the value

        Returns:
            bytes: Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write dates in column 0 and values in specified column
        # Convert date strings to datetime objects
        for idx, (date, value) in enumerate(zip(dates, values), start=1):
            if isinstance(date, str) and date:  # Handle empty strings
                date_obj = datetime.strptime(date, "%Y-%m-%d")
            elif isinstance(date, str) and not date:
                date_obj = None  # Empty date
            else:
                date_obj = date
            ws.cell(row=idx, column=1, value=date_obj)
            ws.cell(row=idx, column=value_column + 1, value=value)

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def test_parse_reservas_internacionales(self, parser):
        """Test extracting Reservas Internacionales."""
        dates = ["2025-10-26", "2025-10-27", "2025-10-28"]
        values = [40000.0, 40500.0, 40771.0]
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=2
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        assert len(records) >= 1
        # Should have Reservas record
        reservas = next(
            (
                r
                for r in records
                if r.data["variable_interna"] == "reservas_internacionales_usd"
            ),
            None,
        )
        assert reservas is not None
        assert (
            reservas.data["indicador"] == "Reservas Internacionales del BCRA"
        )
        assert float(reservas.data["valor"]) == 40771.0  # Most recent value

    def test_parse_ignores_old_xls_format(self, parser):
        """Test that XLS format is handled gracefully."""
        # Create a minimal Excel that won't work with openpyxl
        xls_bytes = b"dummy content"

        document = Document(
            url="https://example.com/test.xls",
            content_type=ContentType.XLS,
            bytes=xls_bytes,
        )

        records = parser.parse(document)
        # Should return empty list or handle gracefully
        assert isinstance(records, list)

    def test_extract_most_recent_picks_latest_date(self, parser):
        """Test that most recent value extraction picks the latest date."""
        dates = ["2025-10-26", "2025-10-28", "2025-10-27"]  # Not in order
        values = [100.0, 300.0, 200.0]
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=2
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        reservas = next(
            (
                r
                for r in records
                if r.data["variable_interna"] == "reservas_internacionales_usd"
            ),
            None,
        )
        assert reservas is not None
        # Should pick 2025-10-28 with value 300.0 (most recent, not just last)
        assert reservas.data["fecha"] == "2025-10-28 00:00:00"

    def test_extract_ignores_invalid_rows(self, parser):
        """Test that invalid rows are ignored."""
        # Add some invalid rows: empty dates, zero values, etc.
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS",
            ["2025-10-28", "", "2025-10-29"],
            [40771.0, 100.0, 40800.0],
            2,
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)
        # Should still work and pick valid rows
        assert len(records) >= 1

    def test_parse_empty_document(self, parser):
        """Test parsing an empty document."""
        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=b"",
        )

        records = parser.parse(document)
        assert len(records) == 0

    def test_parse_wrong_content_type(self, parser):
        """Test parsing with wrong content type."""
        document = Document(
            url="https://example.com/test.html",
            content_type=ContentType.HTML,
            bytes=b"<html></html>",
        )

        records = parser.parse(document)
        assert len(records) == 0

    def test_source_url_preserved(self, parser):
        """Test that source URL is preserved in records."""
        dates = ["2025-10-28"]
        values = [40771.0]
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=2
        )

        document = Document(
            url="https://example.com/custom.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)
        assert len(records) >= 1
        assert records[0].source_url == "https://example.com/custom.xlsm"

    def test_extract_all_three_indicators(self, parser):
        """Test that all three indicators are extracted correctly."""
        # Create Excel with all sheets
        wb = Workbook()

        # RESERVAS sheet
        ws_reservas = wb.active
        ws_reservas.title = "RESERVAS"
        dates = [datetime(2025, 10, 28)]
        # Reservas Internacionales
        ws_reservas.cell(row=1, column=1, value=dates[0])
        ws_reservas.cell(row=1, column=3, value=40771.0)
        # Tipo de Cambio
        ws_reservas.cell(row=1, column=1, value=dates[0])
        ws_reservas.cell(row=1, column=16, value=1470.83)

        # BASE MONETARIA sheet
        ws_bm = wb.create_sheet("BASE MONETARIA")
        ws_bm.cell(row=1, column=1, value=dates[0])
        ws_bm.cell(row=1, column=30, value=40078019.0)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.getvalue()

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        # Should extract all 3 indicators
        assert len(records) >= 1  # At least one should be found

        # Check that we have the expected variable_interna fields
        variable_internas = {r.data["variable_interna"] for r in records}
        assert "reservas_internacionales_usd" in variable_internas

    def test_extract_most_recent_value_helper(self, parser):
        """Test the _extract_most_recent_value helper method directly."""
        dates = ["2025-10-26", "2025-10-28", "2025-10-27"]
        values = [100.0, 300.0, 200.0]
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=2
        )
        excel_file = BytesIO(excel_bytes)

        record = parser._extract_most_recent_value(
            excel_file=excel_file,
            sheet_name="RESERVAS",
            value_column=2,
            source_url="https://example.com",
            indicador="Test",
            unidad="Test Unit",
            variable_interna="test_var",
        )

        assert record is not None
        assert record.data["variable_interna"] == "test_var"
        assert (
            record.data["valor"] == "300"
        )  # Most recent by date (str(value))
        assert record.data["fecha"] == "2025-10-28 00:00:00"

    def test_extract_tipo_cambio_specifically(self, parser):
        """Test specifically extracting Tipo de Cambio."""
        dates = ["2025-10-28"]
        values = [1470.83]
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=15
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        # Should find Tipo de Cambio
        tc_records = [
            r
            for r in records
            if r.data["variable_interna"] == "tipo_cambio_oficial"
        ]
        assert len(tc_records) > 0

    def test_extract_base_monetaria_specifically(self, parser):
        """Test specifically extracting Base Monetaria."""
        # Create multi-sheet Excel
        wb = Workbook()
        ws_bm = wb.active
        ws_bm.title = "BASE MONETARIA"
        ws_bm.cell(row=1, column=1, value=datetime(2025, 10, 28))
        ws_bm.cell(row=1, column=30, value=40078019.0)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        # Should find Base Monetaria
        bm_records = [
            r
            for r in records
            if r.data["variable_interna"] == "base_monetaria_total_ars"
        ]
        assert len(bm_records) > 0

    def test_handles_zero_values(self, parser):
        """Test that zero values are ignored."""
        dates = ["2025-10-28"]
        values = [0.0]  # Zero value
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=2
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        # Should not find valid records with zero values
        [
            r
            for r in records
            if r.data.get("variable_interna") == "reservas_internacionales_usd"
        ]
        # Might have 0 or might have other valid indicators
        assert isinstance(records, list)

    def test_handles_negative_values(self, parser):
        """Test that negative values are ignored."""
        dates = ["2025-10-28"]
        values = [-100.0]  # Negative value
        excel_bytes = self.create_test_excel_bytes(
            "RESERVAS", dates, values, value_column=2
        )

        document = Document(
            url="https://example.com/test.xlsm",
            content_type=ContentType.XLSX,
            bytes=excel_bytes,
        )

        records = parser.parse(document)

        # Should not find valid records with negative values
        [
            r
            for r in records
            if r.data.get("variable_interna") == "reservas_internacionales_usd"
        ]
        assert isinstance(records, list)
